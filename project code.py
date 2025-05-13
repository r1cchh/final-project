import os
import shutil
from datetime import datetime
from openpyxl import Workbook, load_workbook

# folder path (folder which will be sorted)
DOWNLOADS_FOLDER = r"C:\Users\ricar\OneDrive\Dators\TestDownloads"
# log file names
LOG_TEXT_FILE_NAME = "organizer_log.txt"
LOG_EXCEL_FILE_NAME = "organizer_log.xlsx"
# log path
LOG_TEXT_FILE = os.path.join(DOWNLOADS_FOLDER, LOG_TEXT_FILE_NAME)
LOG_EXCEL_FILE = os.path.join(DOWNLOADS_FOLDER, LOG_EXCEL_FILE_NAME)

class FileRecord:
    def __init__(self, name, extension, category, destination):
        self.name = name
        self.extension = extension
        self.category = category
        self.destination = destination
        self.timestamp = datetime.now()

    def to_log_line(self):
        return f"{self.timestamp.strftime('%Y-%m-%d %H:%M:%S')} - Moved: {self.name} -> {self.category}"

    def to_excel_row(self):
        return [
            self.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            self.name,
            self.extension,
            self.category,
            self.destination
        ]
# file type mapping
FILE_CATEGORIES = {
    "Images": [".jpg", ".jpeg", ".png", ".gif", ".bmp"],
    "Documents": [".pdf", ".docx", ".doc", ".txt", ".xlsx", ".pptx"],
    "Videos": [".mp4", ".mov", ".avi", ".mkv"],
    "Music": [".mp3", ".wav", ".flac"],
    "Archives": [".zip", ".rar", ".7z", ".tar", ".gz"],
    "Programs": [".exe", ".msi", ".bat", ".py"],
    "Others": []
}

def get_category(extension):
    for category, extensions in FILE_CATEGORIES.items():
        if extension.lower() in extensions:
            return category
    return "Others"

# check if excel file even exists
def excel_exists():
    if not os.path.exists(LOG_EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Log"
        ws.append(["Date", "Filename", "Extension", "Category", "Destination"])
        wb.save(LOG_EXCEL_FILE)

def log_to_text(record: FileRecord):
    with open(LOG_TEXT_FILE, "a", encoding="utf-8") as f:
        f.write(record.to_log_line() + "\n")

def log_to_excel(record: FileRecord):
    excel_exists()
    wb = load_workbook(LOG_EXCEL_FILE)
    ws = wb.active
    ws.append(record.to_excel_row())
    wb.save(LOG_EXCEL_FILE)

# main 
def organize_folder():
    excel_exists()
    for filename in os.listdir(DOWNLOADS_FOLDER):
        file_path = os.path.join(DOWNLOADS_FOLDER, filename)

        if os.path.isfile(file_path):
            if filename in [LOG_TEXT_FILE_NAME, LOG_EXCEL_FILE_NAME]:
                continue

            _, ext = os.path.splitext(filename)
            category = get_category(ext)

            dest_folder = os.path.join(DOWNLOADS_FOLDER, category)
            os.makedirs(dest_folder, exist_ok=True)

            new_path = os.path.join(dest_folder, filename)
            counter = 1
            base_name, ext_part = os.path.splitext(filename)
            while os.path.exists(new_path):
                new_filename = f"{base_name}({counter}){ext_part}"
                new_path = os.path.join(dest_folder, new_filename)
                counter += 1

            shutil.move(file_path, new_path)
            final_name = os.path.basename(new_path)
            print(f"Moved: {final_name} -> {category}/")

            record = FileRecord(final_name, ext, category, dest_folder)
            log_to_text(record)
            log_to_excel(record)
if __name__ == "__main__":
    organize_folder()
